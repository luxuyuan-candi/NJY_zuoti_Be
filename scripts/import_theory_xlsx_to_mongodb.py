#!/usr/bin/env python3

from __future__ import annotations

import argparse
import hashlib
import os
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pymongo import MongoClient
from pymongo.collection import Collection


QUESTION_HEADERS = [
    "试题GUID",
    "题型(必填）",
    "题干（必填）",
    "答案",
    "选项A",
    "选项B",
    "选项C",
    "选项D",
    "选项E",
    "知识点一级",
    "知识点二级",
    "知识点三级",
    "知识点四级",
    "知识点五级",
    "知识点六级",
    "难度",
    "一致性",
    "分数",
    "解析",
    "出处",
]

LEVEL_CONFIG = {
    "初级理论检查完成": {
        "level": "PRIMARY",
        "label": "初级",
        "practice_set_id": "theory-primary",
        "practice_set_name": "初级理论习题集",
        "mapping_workbook": "初级-医药商品购销员理论知识鉴定要素细目表(1).xlsx",
    },
    "中级理论检查完成": {
        "level": "INTERMEDIATE",
        "label": "中级",
        "practice_set_id": "theory-intermediate",
        "practice_set_name": "中级理论习题集",
        "mapping_workbook": "中级-医药商品购销员理论知识鉴定要素细目表(1).xlsx",
    },
    "高级理论检查完成": {
        "level": "ADVANCED",
        "label": "高级",
        "practice_set_id": "theory-advanced",
        "practice_set_name": "高级理论习题集",
        "mapping_workbook": "高级--医药商品购销员理论知识鉴定要素细目表(1).xlsx",
    },
}

QUESTION_WORKBOOK = "14.理论题库(1).xlsx"
IMPORT_SOURCE = "14.理论题库(1).xlsx"
KNOWLEDGE_COLUMNS = [f"知识点{label}" for label in ("一级", "二级", "三级", "四级", "五级", "六级")]
OPTION_COLUMNS = [f"选项{option}" for option in "ABCDE"]
PUNCTUATION_PATTERN = re.compile(r"[《》“”\"'（）()]")


def clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace("\u3000", " ").strip()
        return text or None
    return value


def squash_text(value: Any) -> str | None:
    cleaned = clean_cell(value)
    if cleaned is None:
        return None
    return re.sub(r"\s+", "", str(cleaned))


def canonical_name(value: Any) -> str | None:
    squashed = squash_text(value)
    if not squashed:
        return None
    return PUNCTUATION_PATTERN.sub("", squashed)


def parse_question_knowledge(value: Any) -> dict[str, str | None] | None:
    raw = clean_cell(value)
    if raw is None:
        return None

    compact = squash_text(raw)
    if not compact:
        return None

    code = None
    name = compact

    wrapped_match = re.match(r"^([A-Z]+)(.+?)\1$", compact)
    if wrapped_match:
        code = wrapped_match.group(1)
        name = wrapped_match.group(2)
    else:
        prefixed_match = re.match(r"^([A-Za-z0-9]+)(.+)$", compact)
        if prefixed_match:
            code = prefixed_match.group(1)
            name = prefixed_match.group(2)

    name = canonical_name(name)
    return {"raw": str(raw), "code": code, "name": name}


def normalize_question_type(value: Any) -> str | None:
    raw = clean_cell(value)
    if raw is None:
        return None

    mapping = {
        "单选题": "single_choice",
        "多选题": "multiple_choice",
        "判断题": "true_false",
    }
    return mapping.get(str(raw), str(raw))


def stable_question_id(practice_set_id: str, row_number: int, stem: str, answer: str, knowledge_leaf: str | None) -> str:
    digest = hashlib.sha1(
        "|".join(
            [
                practice_set_id,
                str(row_number),
                stem,
                answer,
                knowledge_leaf or "",
            ]
        ).encode("utf-8")
    ).hexdigest()
    return f"{practice_set_id}:{digest[:24]}"


@dataclass
class MappingEntry:
    hierarchy_names: list[str]
    point_code: str
    point_name: str
    importance: str | None
    source_row: int


class KnowledgeImportanceIndex:
    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self.sheet_name = "Sheet1"
        self.entries_by_path: dict[tuple[str, ...], MappingEntry] = {}
        self.entries_by_leaf: dict[str, MappingEntry] = {}
        self._load()

    def _load(self) -> None:
        worksheet = load_workbook(self.workbook_path, data_only=True)[self.sheet_name]
        current_names = {
            "一级": None,
            "二级": None,
            "三级": None,
            "四级": None,
        }

        for row_number, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
            values = list(row) + [None] * max(0, 14 - len(row))

            if values[1] is not None:
                current_names["一级"] = canonical_name(values[1])
            if values[3] is not None:
                current_names["二级"] = canonical_name(values[3])
            if values[6] is not None:
                current_names["三级"] = canonical_name(values[6])
            if values[9] is not None:
                current_names["四级"] = canonical_name(values[9])

            point_code = clean_cell(values[11])
            point_name = canonical_name(values[12])
            if not point_code or not point_name:
                continue

            hierarchy_names = [name for name in current_names.values() if name]
            entry = MappingEntry(
                hierarchy_names=hierarchy_names,
                point_code=str(point_code),
                point_name=point_name,
                importance=clean_cell(values[13]),
                source_row=row_number,
            )
            self.entries_by_path[tuple(hierarchy_names + [point_name])] = entry
            self.entries_by_leaf[point_name] = entry

    def match(self, hierarchy_names: list[str]) -> tuple[str | None, str, MappingEntry | None]:
        if not hierarchy_names:
            return None, "unmatched", None

        exact = self.entries_by_path.get(tuple(hierarchy_names))
        if exact:
            return exact.importance, "exact_path", exact

        leaf = hierarchy_names[-1]
        leaf_match = self.entries_by_leaf.get(leaf)
        if leaf_match:
            return leaf_match.importance, "leaf_name", leaf_match

        contains_matches = [
            entry
            for entry_leaf, entry in self.entries_by_leaf.items()
            if entry_leaf in leaf or leaf in entry_leaf
        ]
        if len(contains_matches) == 1:
            match = contains_matches[0]
            return match.importance, "leaf_contains", match

        return None, "unmatched", None


def parse_question_row(
    row_number: int,
    row_values: tuple[Any, ...],
    practice_set_meta: dict[str, str],
    importance_index: KnowledgeImportanceIndex,
) -> dict[str, Any]:
    values = list(row_values[: len(QUESTION_HEADERS)]) + [None] * max(0, len(QUESTION_HEADERS) - len(row_values))
    raw_fields = {
        header: clean_cell(value)
        for header, value in zip(QUESTION_HEADERS, values)
        if clean_cell(value) is not None
    }

    stem = str(raw_fields.get("题干（必填）", "")).strip()
    answer = str(raw_fields.get("答案", "")).strip()
    knowledge_nodes = [parse_question_knowledge(raw_fields.get(column)) for column in KNOWLEDGE_COLUMNS]
    knowledge_nodes = [node for node in knowledge_nodes if node]
    knowledge_path = [node["name"] for node in knowledge_nodes if node.get("name")]
    importance, match_mode, match_entry = importance_index.match(knowledge_path)
    options = [
        {"key": option[-1], "text": raw_fields[option]}
        for option in OPTION_COLUMNS
        if option in raw_fields
    ]
    now = datetime.now(timezone.utc)

    return {
        "_id": stable_question_id(
            practice_set_meta["practice_set_id"],
            row_number,
            stem,
            answer,
            knowledge_path[-1] if knowledge_path else None,
        ),
        "practiceSetId": practice_set_meta["practice_set_id"],
        "practiceSetName": practice_set_meta["practice_set_name"],
        "level": practice_set_meta["level"],
        "levelLabel": practice_set_meta["label"],
        "sourceWorkbook": IMPORT_SOURCE,
        "sourceSheet": practice_set_meta["sheet_name"],
        "sourceRow": row_number,
        "guid": raw_fields.get("试题GUID"),
        "type": normalize_question_type(raw_fields.get("题型(必填）")),
        "typeLabel": raw_fields.get("题型(必填）"),
        "stem": stem,
        "answer": answer,
        "options": options,
        "analysis": raw_fields.get("解析"),
        "origin": raw_fields.get("出处"),
        "difficulty": raw_fields.get("难度"),
        "consistency": raw_fields.get("一致性"),
        "score": raw_fields.get("分数"),
        "knowledge": {
            "level1": raw_fields.get("知识点一级"),
            "level2": raw_fields.get("知识点二级"),
            "level3": raw_fields.get("知识点三级"),
            "level4": raw_fields.get("知识点四级"),
            "level5": raw_fields.get("知识点五级"),
            "level6": raw_fields.get("知识点六级"),
            "pathRaw": [node["raw"] for node in knowledge_nodes],
            "pathNames": knowledge_path,
            "leafCode": knowledge_nodes[-1]["code"] if knowledge_nodes else None,
            "leafName": knowledge_path[-1] if knowledge_path else None,
        },
        "importance": importance,
        "importanceMeta": {
            "matchMode": match_mode,
            "mappingWorkbook": practice_set_meta["mapping_workbook"],
            "mappingSheet": importance_index.sheet_name,
            "mappingRow": match_entry.source_row if match_entry else None,
            "mappingPointCode": match_entry.point_code if match_entry else None,
            "mappingPointName": match_entry.point_name if match_entry else None,
        },
        "rawFields": raw_fields,
        "createdAt": now,
        "updatedAt": now,
    }


def build_practice_set_document(meta: dict[str, str], question_count: int, imported_at: datetime) -> dict[str, Any]:
    return {
        "_id": meta["practice_set_id"],
        "name": meta["practice_set_name"],
        "level": meta["level"],
        "levelLabel": meta["label"],
        "questionCount": question_count,
        "questionSourceWorkbook": IMPORT_SOURCE,
        "sheetName": meta["sheet_name"],
        "mappingWorkbook": meta["mapping_workbook"],
        "status": "ACTIVE",
        "updatedAt": imported_at,
        "createdAt": imported_at,
    }


def load_questions(root_dir: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, dict[str, int]]]:
    workbook = load_workbook(root_dir / QUESTION_WORKBOOK, data_only=True)
    all_questions: list[dict[str, Any]] = []
    practice_sets: list[dict[str, Any]] = []
    stats: dict[str, dict[str, int]] = {}
    imported_at = datetime.now(timezone.utc)

    for sheet_name, config in LEVEL_CONFIG.items():
        meta = {**config, "sheet_name": sheet_name}
        importance_index = KnowledgeImportanceIndex(root_dir / config["mapping_workbook"])
        worksheet = workbook[sheet_name]
        questions: list[dict[str, Any]] = []
        match_counts = {
            "exact_path": 0,
            "leaf_name": 0,
            "leaf_contains": 0,
            "unmatched": 0,
        }

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(clean_cell(value) is not None for value in row[: len(QUESTION_HEADERS)]):
                continue
            question = parse_question_row(row_number, row, meta, importance_index)
            questions.append(question)
            match_counts[question["importanceMeta"]["matchMode"]] += 1

        all_questions.extend(questions)
        practice_sets.append(build_practice_set_document(meta, len(questions), imported_at))
        stats[meta["practice_set_id"]] = match_counts | {"count": len(questions)}

    return practice_sets, all_questions, stats


def replace_documents(collection: Collection, filter_query: dict[str, Any], documents: list[dict[str, Any]]) -> None:
    collection.delete_many(filter_query)
    if documents:
        collection.insert_many(documents, ordered=False)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import theory xlsx questions into MongoDB.")
    parser.add_argument("--root-dir", default="/root/zuoti", help="Directory containing the four xlsx files.")
    parser.add_argument("--mongo-uri", default=os.getenv("MONGO_URI"), help="MongoDB connection URI.")
    parser.add_argument("--mongo-db", default=os.getenv("MONGO_DB", "zuoti_questions"), help="MongoDB database name.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print statistics without writing to MongoDB.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    root_dir = Path(args.root_dir)
    practice_sets, questions, stats = load_questions(root_dir)

    print("Import summary:")
    for practice_set in practice_sets:
        practice_set_id = practice_set["_id"]
        print(
            f"- {practice_set_id}: {practice_set['questionCount']} questions "
            f"(exact={stats[practice_set_id]['exact_path']}, "
            f"leaf={stats[practice_set_id]['leaf_name']}, "
            f"contains={stats[practice_set_id]['leaf_contains']}, "
            f"unmatched={stats[practice_set_id]['unmatched']})"
        )

    if args.dry_run:
        return

    if not args.mongo_uri:
        raise SystemExit("Missing --mongo-uri or MONGO_URI.")

    client = MongoClient(args.mongo_uri)
    db = client[args.mongo_db]
    practice_set_ids = [practice_set["_id"] for practice_set in practice_sets]

    replace_documents(db["practice_sets"], {"_id": {"$in": practice_set_ids}}, practice_sets)
    replace_documents(db["questions"], {"practiceSetId": {"$in": practice_set_ids}}, questions)
    print(f"Wrote {len(practice_sets)} practice sets and {len(questions)} questions into {args.mongo_db}.")


if __name__ == "__main__":
    main()
