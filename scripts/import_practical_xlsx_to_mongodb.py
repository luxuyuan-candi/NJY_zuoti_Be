#!/usr/bin/env python3

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pymongo import MongoClient
from pymongo.collection import Collection


WORKBOOK_CONFIG = {
    "三级试题库.xlsx": {
        "practice_set_id": "practical-primary-level-3",
        "practice_set_name": "三级实操题库",
        "level": "PRIMARY",
        "level_label": "初级",
        "grade_label": "三级",
    },
    "四级试题库.xlsx": {
        "practice_set_id": "practical-intermediate-level-4",
        "practice_set_name": "四级实操题库",
        "level": "INTERMEDIATE",
        "level_label": "中级",
        "grade_label": "四级",
    },
    "五级试题库.xlsx": {
        "practice_set_id": "practical-advanced-level-5",
        "practice_set_name": "五级实操题库",
        "level": "ADVANCED",
        "level_label": "高级",
        "grade_label": "五级",
    },
}

SKIP_SHEETS = {"结构表", "细目表"}
SECTION_PREFIXES = {
    "试题：": "stem",
    "题目：": "stem",
    "考评准备：": "preparation",
    "考评步骤：": "steps",
    "注意：": "notes",
    "考核要点：": "keyPoints",
}
SCORING_HEADER_PREFIX = "序号"
SCORING_HEADER_SECOND = "考核内容"


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\u3000", " ").strip()
    return text or None


def normalize_name(value: str | None) -> str:
    text = clean_text(value) or ""
    text = text.replace("‘", "").replace("’", "").replace("“", "").replace("”", "")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"^[Ss]he", "", text)
    text = text.replace("类用药", "用药")
    text = re.sub(r"\d+$", "", text)
    return text


def row_values_to_list(row: tuple[Any, ...]) -> list[str]:
    return [text for text in (clean_text(cell) for cell in row) if text]


def stable_id(practice_set_id: str, source_sheet: str) -> str:
    digest = hashlib.sha1(f"{practice_set_id}|{source_sheet}".encode("utf-8")).hexdigest()
    return f"{practice_set_id}:{digest[:24]}"


@dataclass
class DetailMapping:
    level1: str | None
    level2: str | None
    leaf_name: str
    importance: str | None
    source_row: int
    remark: str | None

    @property
    def path_names(self) -> list[str]:
        return [part for part in [self.level1, self.level2, self.leaf_name] if part]


def parse_detail_sheet(workbook_path: Path) -> dict[str, DetailMapping]:
    worksheet = load_workbook(workbook_path, data_only=True)["细目表"]
    current_level1 = None
    current_level2 = None
    mappings: dict[str, DetailMapping] = {}

    for row_number, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
        values = list(row) + [None] * max(0, 12 - len(row))
        if clean_text(values[1]):
            current_level1 = clean_text(values[1])
        if clean_text(values[5]):
            current_level2 = clean_text(values[5])
        leaf_name = clean_text(values[8])
        if not leaf_name:
            continue
        mapping = DetailMapping(
            level1=current_level1,
            level2=current_level2,
            leaf_name=leaf_name,
            importance=clean_text(values[9]),
            source_row=row_number,
            remark=clean_text(values[11]),
        )
        mappings[normalize_name(leaf_name)] = mapping
    return mappings


def find_detail_mapping(sheet_name: str, mappings: dict[str, DetailMapping]) -> tuple[DetailMapping | None, str]:
    normalized_sheet = normalize_name(sheet_name)
    if normalized_sheet in mappings:
        return mappings[normalized_sheet], "exact"

    for key, mapping in mappings.items():
        if key and (key in normalized_sheet or normalized_sheet in key):
            return mapping, "contains"
    return None, "unmatched"


def is_scoring_header(row_values: list[str]) -> bool:
    if len(row_values) < 2:
        return False
    return row_values[0] == SCORING_HEADER_PREFIX and SCORING_HEADER_SECOND in row_values[1]


def append_section_value(sections: dict[str, str], section_key: str, text: str) -> None:
    if not text:
        return
    if sections.get(section_key):
        sections[section_key] = f"{sections[section_key]}\n{text}"
    else:
        sections[section_key] = text


def parse_practical_sheet(workbook_name: str, sheet_name: str, worksheet, mapping: DetailMapping | None, match_mode: str, practice_set_meta: dict[str, str]) -> dict[str, Any]:
    rows = [row_values_to_list(row) for row in worksheet.iter_rows(values_only=True)]
    rows = [row for row in rows if row]

    sections = {
        "stem": "",
        "preparation": "",
        "steps": "",
        "notes": "",
        "keyPoints": "",
    }
    current_section = None
    scoring_title = ""
    scoring_rows: list[list[str]] = []
    materials_rows: list[list[str]] = []
    source_rows: list[dict[str, Any]] = []
    scoring_started = False

    for row_number, row in enumerate(rows, start=1):
        source_rows.append({"rowNumber": row_number, "cells": row})

        if scoring_started:
            scoring_rows.append(row)
            continue

        if is_scoring_header(row):
            scoring_started = True
            scoring_rows.append(row)
            current_section = None
            continue

        first_cell = row[0]
        matched_prefix = next((prefix for prefix in SECTION_PREFIXES if first_cell.startswith(prefix)), None)
        if matched_prefix:
            current_section = SECTION_PREFIXES[matched_prefix]
            inline_text = first_cell.removeprefix(matched_prefix).strip()
            append_section_value(sections, current_section, inline_text)
            if len(row) > 1:
                append_section_value(sections, current_section, " | ".join(row[1:]))
            continue

        if "评分标准" in first_cell or first_cell.endswith("标准："):
            scoring_title = first_cell
            current_section = None
            continue

        if len(row) > 1:
            current_section = None
            materials_rows.append(row)
            continue

        if current_section:
            append_section_value(sections, current_section, first_cell)
        elif not scoring_title and "评分标准" in first_cell:
            scoring_title = first_cell
        else:
            materials_rows.append(row)

    stem = sections["stem"] or sheet_name
    path_names = mapping.path_names if mapping else ["未分类", sheet_name]
    leaf_name = mapping.leaf_name if mapping else sheet_name
    now = datetime.now(timezone.utc)

    analysis_parts = [part for part in [sections["keyPoints"], scoring_title] if part]
    return {
        "_id": stable_id(practice_set_meta["practice_set_id"], sheet_name),
        "practiceSetId": practice_set_meta["practice_set_id"],
        "practiceSetName": practice_set_meta["practice_set_name"],
        "level": practice_set_meta["level"],
        "levelLabel": practice_set_meta["level_label"],
        "gradeLabel": practice_set_meta["grade_label"],
        "category": "PRACTICAL",
        "sourceWorkbook": workbook_name,
        "sourceSheet": sheet_name,
        "sourceRow": 1,
        "type": "practical_case",
        "typeLabel": "实操题",
        "stem": stem,
        "answer": "",
        "options": [],
        "analysis": "\n\n".join(analysis_parts) if analysis_parts else None,
        "knowledge": {
            "level1": path_names[0] if len(path_names) > 0 else None,
            "level2": path_names[1] if len(path_names) > 1 else None,
            "level3": path_names[2] if len(path_names) > 2 else None,
            "pathRaw": path_names,
            "pathNames": path_names,
            "leafCode": None,
            "leafName": leaf_name,
        },
        "importance": mapping.importance if mapping else None,
        "content": {
            "preparation": sections["preparation"] or None,
            "steps": sections["steps"] or None,
            "notes": sections["notes"] or None,
            "keyPoints": sections["keyPoints"] or None,
            "materialsRows": materials_rows,
            "scoringTitle": scoring_title or None,
            "scoringRows": scoring_rows,
        },
        "importMeta": {
            "sheetMatchMode": match_mode,
            "detailSourceRow": mapping.source_row if mapping else None,
            "detailRemark": mapping.remark if mapping else None,
        },
        "rawFields": {
            "workbook": workbook_name,
            "sheetName": sheet_name,
            "rows": source_rows,
        },
        "createdAt": now,
        "updatedAt": now,
    }


def build_practical_set_document(meta: dict[str, str], workbook_name: str, question_count: int, chapter_count: int, imported_at: datetime) -> dict[str, Any]:
    return {
        "_id": meta["practice_set_id"],
        "name": meta["practice_set_name"],
        "level": meta["level"],
        "levelLabel": meta["level_label"],
        "gradeLabel": meta["grade_label"],
        "category": "PRACTICAL",
        "questionCount": question_count,
        "chapterCount": chapter_count,
        "sourceWorkbook": workbook_name,
        "status": "ACTIVE",
        "updatedAt": imported_at,
        "createdAt": imported_at,
    }


def load_documents(root_dir: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, dict[str, Any]]]:
    imported_at = datetime.now(timezone.utc)
    all_sets: list[dict[str, Any]] = []
    all_questions: list[dict[str, Any]] = []
    stats: dict[str, dict[str, Any]] = {}

    for workbook_name, meta in WORKBOOK_CONFIG.items():
        workbook_path = root_dir / workbook_name
        workbook = load_workbook(workbook_path, data_only=True)
        detail_mappings = parse_detail_sheet(workbook_path)
        questions = []
        match_counts = {"exact": 0, "contains": 0, "unmatched": 0}
        chapter_names: set[tuple[str, ...]] = set()

        for sheet_name in workbook.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue
            mapping, match_mode = find_detail_mapping(sheet_name, detail_mappings)
            match_counts[match_mode] += 1
            question = parse_practical_sheet(
                workbook_name=workbook_name,
                sheet_name=sheet_name,
                worksheet=workbook[sheet_name],
                mapping=mapping,
                match_mode=match_mode,
                practice_set_meta=meta,
            )
            chapter_path = tuple(question["knowledge"]["pathNames"][:-1] or question["knowledge"]["pathNames"])
            if chapter_path:
                chapter_names.add(chapter_path)
            questions.append(question)

        all_questions.extend(questions)
        all_sets.append(build_practical_set_document(meta, workbook_name, len(questions), len(chapter_names), imported_at))
        stats[meta["practice_set_id"]] = {
            "questionCount": len(questions),
            "chapterCount": len(chapter_names),
            **match_counts,
        }

    return all_sets, all_questions, stats


def replace_documents(collection: Collection, filter_query: dict[str, Any], documents: list[dict[str, Any]]) -> None:
    collection.delete_many(filter_query)
    if documents:
        collection.insert_many(documents, ordered=False)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import practical xlsx question banks into MongoDB.")
    parser.add_argument("--root-dir", default="/tmp/zuoti_practical_import", help="Directory containing the three xlsx files.")
    parser.add_argument("--mongo-uri", default=os.getenv("MONGO_URI"), help="MongoDB connection URI.")
    parser.add_argument("--mongo-db", default=os.getenv("MONGO_DB", "zuoti_questions"), help="MongoDB database name.")
    parser.add_argument("--sets-collection", default="practical_sets", help="Target collection for practice-set metadata.")
    parser.add_argument("--questions-collection", default="practical_questions", help="Target collection for practical questions.")
    parser.add_argument("--dump-dir", default="", help="If set, write parsed JSON files into this directory instead of importing.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print summary without writing to MongoDB.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    root_dir = Path(args.root_dir)
    practice_sets, questions, stats = load_documents(root_dir)

    print("Import summary:")
    for practice_set in practice_sets:
        stat = stats[practice_set["_id"]]
        print(
            f"- {practice_set['_id']}: {stat['questionCount']} sheets, "
            f"{stat['chapterCount']} chapters, "
            f"match(exact={stat['exact']}, contains={stat['contains']}, unmatched={stat['unmatched']})"
        )

    if args.dry_run:
        return

    if args.dump_dir:
        dump_dir = Path(args.dump_dir)
        dump_dir.mkdir(parents=True, exist_ok=True)
        (dump_dir / "practical_sets.json").write_text(
            json.dumps(practice_sets, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        (dump_dir / "practical_questions.json").write_text(
            json.dumps(questions, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        (dump_dir / "practical_import_stats.json").write_text(
            json.dumps(stats, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        print(f"Wrote parsed JSON dumps into {dump_dir}.")
        return

    if not args.mongo_uri:
        raise SystemExit("Missing --mongo-uri or MONGO_URI.")

    client = MongoClient(args.mongo_uri)
    db = client[args.mongo_db]
    practice_set_ids = [item["_id"] for item in practice_sets]

    replace_documents(db[args.sets_collection], {"_id": {"$in": practice_set_ids}}, practice_sets)
    replace_documents(db[args.questions_collection], {"practiceSetId": {"$in": practice_set_ids}}, questions)
    print(
        f"Wrote {len(practice_sets)} practical sets into {args.sets_collection} "
        f"and {len(questions)} practical questions into {args.questions_collection}."
    )


if __name__ == "__main__":
    main()
